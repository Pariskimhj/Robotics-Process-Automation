from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1) # 열 한 칸 옮기기
ws['B1'].value = "국어" # B1 셀에 '국어' 입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1) # 밑으로 5줄, 왼쪽으로 1칸 이동 (기존에 있던 데이터는 덮어쓰기)


wb.save("sample_korean.xlsx")