from openpyxl import load_workbook
wb = load_workbook('sample.xlsx') # sample 파일 불러오기
ws = wb.active

for row in ws.iter_rows(min_row=2): # 두번째 row부터 전체
	# 번호, 영어, 수학
	if int(row[1].value) > 80: # 영어 점수가 80점 초과면
		print(row[0].value, "번 학생은 영어 천재") # (번호)번 학생은 영어 천재 출력

for row in ws.iter_rows(max_row=1): # 첫번째 row만
	for cell in row: # row의 cell
		if cell.value == '영어': # cell의 값이 영어면
			cell.value = '컴퓨터' # cell의 값을 컴퓨터로 바꾸기

wb.save('sample_modified.xlsx') # 저장