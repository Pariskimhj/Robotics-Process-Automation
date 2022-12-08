from types import new_class
from openpyxl import Workbook
wb = Workbook()
wb.active
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
wb.save('sample.xlsx') # sample.xlsx로 저장

ws.title = 'Mysheet'   # Sheet 이름 변경

ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경 
# 구글 RBG Caculator :  https://www.w3schools.com/colors/colors_rgb.asp

ws1 = wb.create_sheet('YourSheet') # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet('NewSheet', 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test" # NewSheet A1에 Test 입력
target = wb.copy_worksheet(new_ws) # wb의 new_ws sheet를 복사해서 target에 넣는다.
target.title = 'Copied Sheet' # target의 sheet name을 Copied Sheet로 변경

wb.save('sample.xlsx') # 파일 저장