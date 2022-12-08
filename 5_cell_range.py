from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 데이터 1줄씩 데이터 넣기

ws.append(['번호', '영어', '수학'])
for i in range(1, 11): # 10개 데이터 넣기
	ws.append([i, randint(0, 100), randint(0, 100)])

# 특정 column 가져오기

col_B = ws["B"] # 영어 column만 가져오기
print(col_B) # col_B에 대한 정보 출력
for cell in col_B: # col_B의 cell
	print(cell.value) # cell의 값 출력

col_range = ws['B:C'] # 영어, 수학 column 함께 가지고 오기
for cols in col_range:
	for cell in cols:
		print(cell.value)

# 특정 row 가져오기

row_title = ws[1] # 첫번째 row만 가지고 오기
for cell in row_title:
	print(cell.value)

row_range = ws[2:6] # title 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
for rows in row_range:
	for cell in rows:
		print(cell.value, end=' ')
	print()

from openpyxl.utils.cell import coordinate_from_string # cell의 좌표 정보

# 좌표 정보 불러오기

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
	for cell in rows:
		print(cell.value, end=" ")
		print(cell.coordinate, end=" ") # cell의 좌표 정보 / A1, A2, B1, B2 ...
		xy = coordinate_from_string(cell.coordinate) # cell의 좌표 정보 / ('A', 2) ('B', 2) ('C', 2)...
		print(xy, end=' ')
		print(xy[0], end='') # A
		print(xy[1], end=' ') # 1
	print()

#전체 rows 값 조회

print(tuple(ws.rows))
for row in tuple(ws.rows):
	print(row)
	print(row[2].value) # index 2(B열)의 value

#전체 columns 값 조회

print(tuple(ws.columns))
for column in tuple(ws.columns):
	print(column[0].value)# index 0(1행)의 value

#전체 rows 값 조회

for row in ws.iter_rows(): # 전체 row
	print(row[1].value)

#전체 columns 값 조회

for column in ws.iter_cols(): # 전체 column
	print(column[0].value) # 1행 전체 출력


# 특정 범위 가져오기

for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): #1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
	print(row[0].value, row[1].value) # 수학, 영어
	print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
	print(col)

wb.save('sample.xlsx')