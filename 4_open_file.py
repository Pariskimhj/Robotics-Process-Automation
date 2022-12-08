
from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook('sample.xlsx') # sample.xlsx 파일을 wb로 불러오기
ws = wb.active # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=' ') # cell 값 불러오기, cell 사이는 띄어쓰기
    print() # 행 끝나면 한 줄 Enter

# cell 갯수를 모를 때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=' ') # cell 값 불러오기, cell 사이는 띄어쓰기
    print() # 행 끝나면 한 줄 Enter